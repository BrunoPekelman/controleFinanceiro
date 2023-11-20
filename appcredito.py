import re
import pandas as pd
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


r = r'\d{2}.\d{2}.\d{4}'

#definir arquivo-fatura
fatura_txt = "FaturaBB.txt"

#def arquivo final excel
arquivo_final = 'Finanças.xlsx'
planilha = 'gastos'

#definir lista temporaria
lista_temp = []

# definir lista final
lista_final = []

#def lista parcelas
parcelas = []

#ler fatura 
with open(fatura_txt, "r") as arquivo_txt:
    linhas = arquivo_txt.readlines()

#arrumar linhas sem info
for linha in linhas:
    linha = linha.strip()
    
#separar linhas filtradas na variavel r (datas)
for linha in linhas:
    if re.match(r,linha):
        #apendar linhas na lista
        lista_temp.append(linha)

lista_temp = lista_temp[1:]

#inicio criação do Data frame
colunas = ['Data', 'Descrição', 'Parcela', 'Cidade', 'País', 'Valor', 'Tipo']
df = pd.DataFrame(columns=colunas)

palavras_chave = {
    ('ifood', 'rappi', 'RAPPI*MIMIC FORNECIMENT', 'DL *Rappi BR  Prime Pr'): 'delivery',
    ('veloe', 'ipiranga', 'abastece', 'abasteceai'): 'carro',
    ('AmazonPrimeBR', 'Microsoft*Ultimate 1 Mo', 'DL*GOOGLE YouTub'): 'Assinatura',
    ('IOF - COMPRA NO EXTERIO'): 'IOF',
}


for i in lista_temp:
    data = i[0:10]
    descr = i[10:33]
    cidade = i[33:46]
    pais = i[47:49]
    valor = i[60:71]
    parcelas = '1x'
    tipo = ""
    if 'PARC' in descr:
        descr = i[10:24]
        cidade = i[35:45]
        parcelas = i[29:35]
        tipo = ""
    nova_linha = pd.Series([data, descr, parcelas, cidade, pais, valor, tipo], index=colunas)
    df = df._append(nova_linha, ignore_index=True)



df['Valor'] = df['Valor'].str.replace('.', '').str.replace(',', '.')
df['Valor'] = df['Valor'].astype(float)

def categorizar_descricao(descricao):
    for palavras, tipo in palavras_chave.items():
        if any(palavra.lower() in descricao.lower() for palavra in palavras):
            print(f'Correspondência encontrada: {palavras} - {tipo}')
            return tipo
    
    print(f'Nenhuma correspondência encontrada para: {descricao}')
    return 'Outros'  # Se nenhuma correspondência for encontrada, categorize como 'Outro'

# Aplicar a função à coluna 'Descrição' para criar a coluna 'Tipo'
df['Tipo'] = df['Descrição'].apply(categorizar_descricao)




df_final = pd.DataFrame()
df_final = df.loc[df['Valor'] >= 0]
 


df_final = df_final.reset_index(drop=True)


print(df_final)

book = openpyxl.load_workbook(arquivo_final)

# Escolha a planilha
sheet = book[planilha]

# Encontre a última linha com dados na planilha
row = sheet.max_row + 1

# Adicione os dados do df_final começando da última linha
for record in dataframe_to_rows(df_final, index=False, header=False):
    for col, val in enumerate(record, start=1):
        sheet.cell(row=row, column=col, value=val)
    row += 1

# Salve as alterações
book.save(arquivo_final)










        

        
