import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ProcessadorFatura:
    def __init__(self, fatura_txt, arquivo_final, planilha):
        self.fatura_txt = fatura_txt
        self.arquivo_final = arquivo_final
        self.planilha = planilha
        self.lista_temp = []

    def filtrar_linhas(self, linhas):
        for linha in linhas:
            linha = linha.strip()
            if re.match(r'\d{2}.\d{2}.\d{4}', linha):
                self.lista_temp.append(linha)

        self.lista_temp = self.lista_temp[1:]

    def criar_dataframe(self):
        colunas = ['Data', 'Descrição', 'Parcela', 'Cidade', 'País', 'Valor', 'Tipo']
        df = pd.DataFrame(columns=colunas)

        for i in self.lista_temp:
            data = i[0:10]
            descr = i[10:33]
            cidade = i[33:46]
            pais = i[47:49]
            valor = i[60:71]
            parcelas = '1x'
            tipo = ""

            if 'PARC' in descr:
                descr = i[10:24]
                cidade = i[35:45]
                parcelas = i[29:35]
                tipo = ""

            nova_linha = pd.Series([data, descr, parcelas, cidade, pais, valor, tipo], index=colunas)
            df = df._append(nova_linha, ignore_index=True)

        df['Valor'] = df['Valor'].str.replace('.', '').str.replace(',', '.')
        df['Valor'] = df['Valor'].astype(float)

        return df

    def categorizar_descricao(self, descricao):
        palavras_chave = {
            ('ifood', 'rappi', 'RAPPI*MIMIC FORNECIMENT', 'DL *Rappi BR  Prime Pr', 'RAPPI*RAPPI BRASIL'): 'delivery',
            ('veloe', 'ipiranga', 'abastece', 'abasteceai', 'automo'): 'carro',
            ('AmazonPrimeBR', 'Microsoft*Ultimate 1 Mo', 'DL*GOOGLE YouTub', 'kaspersky'): 'Assinatura',
            ('iof',): 'IOF  ',
            ('Amazon Prime Canais',): 'Compras Online',
        }

        for palavras, tipo in palavras_chave.items():
            if any(palavra.lower() in descricao.lower() for palavra in palavras):
                print(f'Correspondência encontrada: {palavras} - {tipo} para a descrição: {descricao}')
                return tipo

        print(f'Nenhuma correspondência encontrada para: {descricao}')
        return 'Outros'

    def processar_fatura(self):
        with open(self.fatura_txt, "r") as arquivo_txt:
            linhas = arquivo_txt.readlines()

        self.filtrar_linhas(linhas)
        df = self.criar_dataframe()

        # Aplicar a função à coluna 'Descrição' para criar a coluna 'Tipo'
        df['Tipo'] = df['Descrição'].apply(self.categorizar_descricao)

        df_final = df.loc[df['Valor'] >= 0]
        df_final = df_final.reset_index(drop=True)

        return df_final

    def exportar_excel(self, df_final):
        book = load_workbook(self.arquivo_final)

        # Escolha a planilha
        sheet = book[self.planilha]

        # Encontre a última linha com dados na planilha
        row = sheet.max_row + 1

        # Adicione os dados do df_final começando da última linha
        for record in dataframe_to_rows(df_final, index=False, header=False):
            for col, val in enumerate(record, start=1):
                sheet.cell(row=row, column=col, value=val)
            row += 1

        # Salve as alterações
        book.save(self.arquivo_final)

# Exemplo de uso
fatura_processor = ProcessadorFatura(fatura_txt="fatura1123.txt", arquivo_final="Finanças.xlsx", planilha="gastos")
df_final = fatura_processor.processar_fatura()
fatura_processor.exportar_excel(df_final)
print(df_final)